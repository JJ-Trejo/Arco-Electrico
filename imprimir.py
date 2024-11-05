import numpy as np
import openpyxl #Importa la biblioteca Openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
import pandas as pd

def imprime (system_data, res):

    # Cargar el archivo existente en lugar de crear uno nuevo
    # try:
    #     wb = openpyxl.load_workbook("Calculos_IEEE_1584_2018.xlsx") # Abrir el archivo existente
    # except FileNotFoundError:
    #     wb = openpyxl.Workbook() # Si el archivo no existe, crear uno nuevo    
    
    wb = openpyxl.Workbook() # Si el archivo no existe, crear uno nuevo   
    ws = wb.active #Seleccionar la hoja activa (por defecto se crea una hoja)
    ws.title = "Resultados" #Cambiamos el nombre de la hoja que se genera
    ws['A1'] = 'DATOS DEL SISTEMA'
    ws['A1'].font = Font(bold=True, size=12) ## NEGRITA
    ws['A12'] = 'RESULTADOS'
    ws['A12'].font = Font(bold=True, size=12) ## NEGRITA

    ws['A13'] = 'CORRIENTE NORMAL'
    ws['A13'].font = Font(bold=True, size=12) ## NEGRITA
    ws['A14'] = 'Variable'
    ws['A15'] = 'I_arc'
    ws['A16'] = 'T'
    ws['A17'] = 'E'
    ws['A18'] = 'AFB'

    ws['B14'] = 'Unidades'
    ws['B15'] = 'kA'
    ws['B16'] = 'ms'
    ws['B17'] = 'J/cm2'
    ws['B18'] = 'mm'

    ws['C14'] = 'Valor'
    ws['C15'] = res[0][0]
    ws['C16'] = res[1][0]
    ws['C17'] = res[2][0]
    ws['C18'] = res[3][0]

    ws['E13'] = 'CORRIENTE REDUCIDA'
    ws['E13'].font = Font(bold=True, size=12) ## NEGRITA
    ws['E14'] = 'Variable'
    ws['E15'] = 'I_arc'
    ws['E16'] = 'T'
    ws['E17'] = 'E'
    ws['E18'] = 'AFB'

    ws['F14'] = 'Unidades'
    ws['F15'] = 'kA'
    ws['F16'] = 'ms'
    ws['F17'] = 'J/cm2'
    ws['F18'] = 'mm'

    ws['G14'] = 'Valor'
    ws['G15'] = res[0][1]
    ws['G16'] = res[1][1]
    ws['G17'] = res[2][1]
    ws['G18'] = res[3][1]

    if (res[2][0] > res[2][1]):
        ws['C17'].font = Font(color = 'FF0000',bold=True, size=12) ## red
    else:
        ws['G17'].font = Font(color = 'FF0000',bold=True, size=12) ## red

    if (res[3][0] > res[3][1]):
        ws['C18'].font = Font(color = 'FF0000',bold=True, size=12) ## red
    else:
        ws['G18'].font = Font(color = 'FF0000',bold=True, size=12) ## red

    wb.save("Calculos_IEEE_1584_2018.xlsx") #Guardamos el archivo

    #Escribir el DataFrame en un archivo de Excel
    #Sheet_name: selecciona la hoja donde se va a escribir NOTE: ESTA YA TIENE QUE EXISTIR
    # Usa ExcelWriter en modo "append" (a) para no sobrescribir el archivo
    with pd.ExcelWriter('Calculos_IEEE_1584_2018.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        #df: Data Frame | columns indica los titulos de las columnas 1, 2 y 3
        df_1 = pd.DataFrame(system_data, columns = ['VARIABLE', 'UNIDADES', 'VALOR'])
        # Escribir el DataFrame debajo de "DATOS DEL SISTEMA"
        df_1.to_excel(writer, startrow=1, startcol=0, sheet_name="Resultados", index=False) # index=False evita que se guarde el índice numérico de pandas
    
